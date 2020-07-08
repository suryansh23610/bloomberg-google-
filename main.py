from openpyxl import Workbook 
from openpyxl import load_workbook
import requests
from lxml import html
import argparse
from selenium import webdriver
import os
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
import time
from lxml.html.clean import Cleaner
import configparser
import logging
from csv import writer
from csv import reader


##########################################################################################################################################
##### Function to get the webpage content in  a string format ############################################################################
def get_string(url) :

    headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64;     x64; rv:66.0) Gecko/20100101 Firefox/66.0", "Accept-Encoding":"gzip, deflate",     "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8", "DNT":"1","Connection":"close", "Upgrade-Insecure-Requests":"1"}

    cleaner = Cleaner()
    cleaner.javascript = True 
    cleaner.style = True
    response=requests.get(url,headers=headers,timeout=10)
    parser=html.fromstring(cleaner.clean_html(response.text))
    l=[e.strip() for e in parser.xpath('//body//div//text()')]
    l=' '.join(l)
    return l


##########################################################################################################################################
###### Function to get data from boolmberg site otherwise  return none ###################################################################
def bloomberg_data(custom_query,dr):

    headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64;     x64; rv:66.0) Gecko/20100101 Firefox/66.0", "Accept-Encoding":"gzip, deflate",     "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8", "DNT":"1","Connection":"close", "Upgrade-Insecure-Requests":"1"}
    try: ### Get data from google
        dr.get("http://www.google.com")
        print('Opened google...')
        #time.sleep(1)
    except Exception as open_google:
        print('Exception in opening Google')
        print(type(open_google))
        print(str(open_google))
    try:
        que=dr.find_element_by_xpath("//input[@name='q']")
        que.send_keys(custom_query)
        que.send_keys(Keys.RETURN)
        time.sleep(2)
        link_web=dr.find_elements_by_xpath('//div[@class="r"]')
        link_web=[ e.find_element_by_xpath('.//a') for e in link_web]
        links=[e.get_attribute('href') for e in link_web]
        for link in links:
            if 'bloomberg' in link.lower():
                response=requests.get(link,headers=headers)
                parser=html.fromstring(response.text)
                l=parser.xpath('//div[@class="infoTableItemValue__e188b0cb"]//text()')
                return l[1],l[2]
        return None
    except Exception as bloom_er:
        #print(bloom_er)
        return None
##########################################################################################################################################


##########################################################################################################################################
##### function to write data in the excel ###############################################################################################
def write_to_excel(wb,Sheet,row_,s0,s1,s2,s3,s4,s5,path, max_col):
    Sheet.cell(row=row_,column=max_col+1).value=s0
    Sheet.cell(row=row_,column=max_col+2).value=s1
    Sheet.cell(row=row_,column=max_col+3).value=s2
    Sheet.cell(row=row_,column=max_col+4).value=s3
    Sheet.cell(row=row_,column=max_col+5).value=s4
    Sheet.cell(row=row_,column=max_col+6).value=s5
    wb.save(path)
##########################################################################################################################################


##########################################################################################################################################
### Function to get data from google website otherwise get the data from the first three links   #########################################
def _main(custom_query,dr,row_):
    try: ### Get data from google
        dr.get("http://www.google.com")
        print('Opened google...')
        #time.sleep(5)
    except Exception as open_google:
        print('Exception in opening Google')
        print(type(open_google))
        print(str(open_google))

    que=dr.find_element_by_xpath("//input[@name='q']")
    que.send_keys(custom_query)
    que.send_keys(Keys.RETURN)
    time.sleep(1)
    try:
        website=dr.find_element_by_xpath('//a[@class="ab_button"]').get_attribute('href')
        print(website)
        website_data=get_string(website)
        string1,string2,string3='','',''
    except Exception as e:
        #print(e)
        link_web=dr.find_elements_by_xpath('//div[@class="r"]')[:3]
        link_web=[ e.find_element_by_xpath('.//a') for e in link_web]
        links=[e.get_attribute('href') for e in link_web]
        print(links)
        try:
            string1=get_string(links[0])
        except requests.exceptions.Timeout as link1_err :
            link1_err=f'At row:- {row_},query = {custom_query} , error:- {str(link1_err)}'
            logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
            logging.warning(link1_err)
            string1=''
        try:
            string2=get_string(links[1])
        except requests.exceptions.Timeout as link2_err :
            link2_err=f'At row:- {row_},query = {custom_query} , error:- {str(link2_err)}'
            logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
            logging.warning(link2_err)
            string2=''
        try:
            string3=get_string(links[2])
        except requests.exceptions.Timeout as link3_err :
            link3_err=f'At row:- {row_},query = {custom_query} , error:- {str(link3_err)}'
            logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
            logging.warning(link3_err)
            string3=''
        website_data=''
    return website_data,string1,string2,string3
##########################################################################################################################################



if __name__ == '__main__':
    ##############################################################
    ### Reading configuration file
    config = configparser.ConfigParser()
    config.read(r'config.ini')
    path = config['file_path']['path']
    row_= int(config['file_path']['row'])
    columns=config.get('columns', 'cols')
    required_column=columns.split(',')
    
    ##############################################################
    ### this will check if the file contains .csv extension or not
    CSV=False
    if '.csv' in path.lower():
        seperator=config['csv_seperator']['csv_seperator']
        print(seperator)
        CSV=True
    ##############################################################


    ##############################################################
    #### initialise the selenium chrome driver ###################
    chrome_options=Options()
    chrome_options.add_argument("--headless")
    dr = webdriver.Chrome(options=chrome_options,executable_path=r"chromedriver.exe")
    time.sleep(3)
    ##############################################################

    ##############################################################
    #### this section will execute if excel file is given as input
    if not CSV:
        wb = load_workbook(path)
        sheets = wb.sheetnames
        Sheet = wb[sheets[0]]
        max_r=Sheet.max_row

        column_names={}
        Current  = 1
        for COL in Sheet.iter_cols(1, Sheet.max_column):
            column_names[COL[0].value] = Current
            Current += 1

        #### to make sure writing of excel file takes place from first row onward only
        if row_<2:
            row_=2
       
        max_col=Sheet.max_column
        l=[]
        for i in range(1,max_col+1):
            data=Sheet.cell(row=1,column=i).value
            if data is not None:
                l.append(Sheet.cell(row=1,column=i).value)

        new=True
        for e in l:
            if e=='Bloomberg Data':
                max_col-=6
                new=False
        if new:
            Sheet.cell(row=1,column=max_col+1).value='Bloomberg Data'
            Sheet.cell(row=1,column=max_col+2).value='Google Data'
            Sheet.cell(row=1,column=max_col+3).value='Link1 Data'
            Sheet.cell(row=1,column=max_col+4).value='Link2 Data'
            Sheet.cell(row=1,column=max_col+5).value='Link3 Data'
            Sheet.cell(row=1,column=max_col+6).value='Final Data'
        
        

        ##### this loop will run till data till last row is written
        limit=1
        while row_<=max_r:
            try:
                custom_query=''
                for e in required_column:
                    cell_value=Sheet.cell(row=row_,column=column_names[e]).value
                    if cell_value is not None:
                        custom_query+=str(cell_value)+','
                custom_query=custom_query[:-1]
                

                #### custom_query= it is the str which will be searched in the google 

                name=custom_query.split(',')[0]
                print(custom_query)
                print(name)
                data=bloomberg_data(name,dr)
                print(data)
                if data is None:
                    s1,s2,s3,s4=_main(custom_query,dr,row_)
                    s0=''

                else:
                    s0=','.join(data)
                    s1,s2,s3,s4='','','',''
                s5=' '.join([s0,s1,s2,s3,s4])
                
                write_to_excel(wb,Sheet,row_,s0,s1,s2,s3,s4,s5,path, max_col)
                row_+=1
                limit=1
            except Exception as write_err:
                ### write logging in the log file 
                write_err=f'At row:- {row_},query = {custom_query} , error:- {str(write_err)}'
                logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
                logging.warning(write_err)
                if limit>=2:
                    row_+=1
                    limit=1
                if 'xpath' not in str(write_err):
                    limit+=1
                time.sleep(3)

            ### it handles the keyboard interrupt exception ######################################################
            except KeyboardInterrupt:
                logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
                logging.warning(f'At row:- {row_} KeyboardInterrupted')
                wb.save(path)
                break

    ##################################################################################################################
    ####### this section will execute when csv file is given as input ###############################################
    else:
        with open(path, 'r') as read_obj,open('output_2.csv', 'w', newline='') as write_obj:

            # Create a csv.reader object from the input file object
            csv_reader = reader(read_obj,delimiter=seperator)

            # Create a csv.writer object from the output file object
            csv_writer = writer(write_obj,delimiter=seperator)
            i=0

            ### Read each row of the input csv file as list
            for row in csv_reader:
                if i==0:
                    d={}
                    for (i,v) in enumerate(row):
                        d[v]=i
                    row.extend(['Bloomberg Data','Google Data','Link1 Data','Link2 Data','Link3 Data','Final Data'])
                    csv_writer.writerow(row)
                else:
                    error=True
                    custom_query=''
                    for e in required_column:
                        custom_query+=str(row[d[e]])+','
                    name=custom_query.split(',')[0]
                    print(custom_query)
                    print(name)

                    ##  this loop will handle the exception
                    while error:
                        try:
                            data=bloomberg_data(name,dr)
                            print(data)
                            if data is None:
                                s1,s2,s3,s4=_main(custom_query,dr)
                                s0=''

                            else:
                                s0=','.join(data)
                                s1,s2,s3,s4='','','',''
                            s5=' '.join([s0,s1,s2,s3,s4])
                            l=[s0,s1,s2,s3,s4,s5]
                            row.extend(l)
                            csv_writer.writerow(row)
                            error=False
                        except Exception as csv_err:
                            csv_err=f'At row:- {i},query = {custom_query} , error:- {str(csv_err)}'

                            ###### write the logging for error ############################################
                            logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
                            logging.warning(csv_err)
                            time.sleep(5)
                        except KeyboardInterrupt:
                            logging.basicConfig(filename='app.log', filemode='w', format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
                            logging.warning(f'At row:- {i} KeyboardInterrupted')
                            wb.save(path)
                            break
                i+=1
        ### get the absolute name of the file from the path given
        new_file_name=os.path.basename(path)
        try:
            ### rename the output file as input one ##
            os.rename('output_2.csv',new_file_name)
        except:
            ### this section will exceute when input file is in same folder as script
            os.remove(path)
            os.rename('output_2.csv',new_file_name)
    #########################################################################################################################